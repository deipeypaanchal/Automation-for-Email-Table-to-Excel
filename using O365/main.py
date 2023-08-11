"""

This pragram was written by Deepey Panchal and Brandon Pyle

IMPORTANT INSTRUCTIONS:

When you run the program, a URL will be printed to the terminal. Go to the link and sign in with your exelon account
If your login attempt is successful, the page will refresh. Once it refreshes, copy the URL in the search bar and paste
it into the terminal. Your account should be successfully authenticated and the program will run.

Change the MASTER_WORKBOOK_PATH variable to the name of your excel file.
Change the EMAIL_FOLDER_NAME variable to the name of the folder where your emails will be stored.

"""

import schedule
import time
import openpyxl
import pandas as pd
import O365
import logging

CLIENT_ID = '' # Client ID from Azure Application
CLIENT_SECRET = '' # Client Secret Value from Azure Application
MASTER_WORKBOOK_PATH = '' # Path to the master Excel workbook
EMAIL_FOLDER_NAME = '' # Name of the folder where your emails will be located

class EquipmentEntry:
    def __init__(self, site, equipment, message, last_state_change):
        self.site = site
        self.equipment = equipment
        self.message = message
        self.last_state_change = last_state_change

# Function to authenticate the user
def authenticate_account():
    account = O365.Account(credentials=(CLIENT_ID, CLIENT_SECRET), scopes=['basic', 'mailbox']) # Authenticates the user and gives basic and mailbox access to this program
    account.authenticate() # Creates the account variable to be used in accessing the user's mailbox

    return account

# Function to convert HTML data to a format excel accepts
def parse_html_email(html_content, worksheet):
    table = html_content.find('table') # Finds all tables in the HTML and adds them to the table variable

    # If no tables exist, exit the function and return None
    if not table:
        return None
    # Otherwise, do the following
    else:
        table_data = [] # List variable to hold the table data

        if worksheet == 1:
            # Appends each row of the HTML table to the table_data variable
            for row in table.find_all('tr'):
                row_data = [cell.text.strip() for cell in row.find_all('td')]
                if row_data:
                    table_data.append(row_data)
        else:
            # Appends each row with column headers to the table_data variable
            for row in table.find_all('tr')[1:]:
                columns = row.find_all('td')
                site, equipment, message, last_state_change = (col.text.strip() for col in columns)
                table_data.append(EquipmentEntry(site, equipment, message, last_state_change))

        return table_data

# Function to append table data to sheet 1 of the worksheet
def append_table_data_to_worksheet1(table_data, worksheet1, added_rows):
    if table_data:
        for row_data in table_data[1:]:
            site, equipment, message, last_state_change = row_data[0], row_data[1], row_data[2], row_data[3]
            if (site, equipment, message, last_state_change) not in added_rows:
                added_rows.add((site, equipment, message, last_state_change))
                # Check if any of the values in the specific columns are not in the worksheet
                if not any(
                    (site, equipment, message, last_state_change) == (ws_site, ws_equipment, ws_message, ws_last_state_change)
                    for ws_site, ws_equipment, ws_message, ws_last_state_change in worksheet1.iter_rows(
                        values_only=True, min_col=1, max_col=4
                    )
                ):
                    worksheet1.append(row_data)

# Function to update the sheet 3 with new data
def update_table_data(existing_entries, new_entries, worksheet3):
    for new_entry in new_entries:
        for existing_entry in existing_entries:
            if new_entry.equipment == existing_entry.equipment and new_entry.site == existing_entry.site:
                existing_entry.message = new_entry.message
                existing_entry.last_state_change = new_entry.last_state_change
                break
        else:
            existing_entries.append(new_entry)

    # After updating the data in existing_entries, write it back to worksheet3
    worksheet3.delete_rows(2, worksheet3.max_row)  # Delete existing data (excluding header row)

    # Write the updated data to worksheet3
    for entry in existing_entries:
        worksheet3.append([entry.site, entry.equipment, entry.message, entry.last_state_change])

# Function to get current data in list form from sheet 3
def read_existing_data_from_excel(worksheet3):
    try:
        data = list(worksheet3.values)  # Convert the worksheet data to a list of lists
        df = pd.DataFrame(data[1:], columns=data[0])  # Skip the header row
        existing_entries = [EquipmentEntry(row["Site"], row["Equipment"], row["Message"], row["Last State Change"]) for _, row in df.iterrows()]
    except:
        existing_entries = []

    return existing_entries

def driver(account):
    loopSuccess = False

    try:
        mailbox = account.mailbox() # Connects to the accounts mailbox
        inbox = mailbox.get_folder(folder_name=EMAIL_FOLDER_NAME) # Sets the inbox variable to the folder selected in quotes
        emails = inbox.get_messages(limit=100) # Gets up to 100 emails from this folder

        try:
            # Open the master workbook and access the sheets
            workbook = openpyxl.load_workbook(MASTER_WORKBOOK_PATH)

            worksheet1 = workbook["Sheet1"]
            worksheet2 = workbook["Sheet2"]
            worksheet3 = workbook["Sheet3"]

            print("\nSuccessfully opened the excel workbook\n")

            # Keep track of added rows to avoid duplicates using a set
            added_rows = set(tuple(row) for row in worksheet1.iter_rows(values_only=True))

            # Creates a list to store existing data from worksheet 3
            existing_entries = read_existing_data_from_excel(worksheet3)

            # Loops through each email in the inbox from the selected sender
            for email in emails:
                loopSuccess = True

                print(f"Looping through '{email.subject}' from the {inbox.name} folder and adding data to the excel workbook...\n")

                # Gets the HTML content of the current email
                data = email.get_body_soup()

                # Extract the table data from the HTML body and add it to sheet 1
                table_data = parse_html_email(data, 1)

                # Append the table data to worksheet1, excluding duplicates
                append_table_data_to_worksheet1(table_data, worksheet1, added_rows)

                # Extract the table data from the HTML body and add it to sheet 3
                new_entries = parse_html_email(data, 3)
                if new_entries:
                    update_table_data(existing_entries, new_entries, worksheet3)

                if table_data is None or new_entries is None:
                    print("ERROR: No tables found in the email\n")
                else:
                    print("Data added to Worksheet 1.\n")

                # Convert the combined data to a Pandas DataFrame
                data_dict = {
                    "Site": [entry.site for entry in existing_entries],
                    "Equipment": [entry.equipment for entry in existing_entries],
                    "Message": [entry.message for entry in existing_entries],
                    "Last State Change": [entry.last_state_change for entry in existing_entries]
                }
                df = pd.DataFrame(data_dict)

                # Save the data directly to the "Sheet3" of the master workbook using Openpyxl
                worksheet3.delete_rows(2, worksheet3.max_row)  # Delete existing data (excluding header row)
                for _, row in df.iterrows():
                    worksheet3.append(row.tolist())

                print("Database in worksheet 3 is updated.")

                # Find all rows with "not working" in the "Message" column
                not_working_rows = set()
                for row in worksheet3.iter_rows(values_only=True):
                    if row[2] == "not working":  # Assuming "Message" is the 3rd column (index 2)
                        not_working_rows.add(tuple(row))

                # Clear existing data in Sheet2
                worksheet2.delete_rows(2, worksheet2.max_row)

                # Append the not_working_rows to Sheet2
                for row_data in not_working_rows:
                    worksheet2.append(row_data)
                
                print("\nWorksheet 2 updated with inactive devices.")
                print("\nTable data added/updated to the master worksheet.\n\n")
        except Exception as e:
            print("\n\nERROR: There was an error opening the excel file. Please make sure the file name is correct in the code and the file is not open when the code is executed.")
            logging.exception(e)

    except Exception as e:
        print("\n\nERROR: An error has occurred")
        logging.exception(e)

    # Save the modified workbook
    try:
        workbook.save(MASTER_WORKBOOK_PATH)
        workbook.close()
    except Exception as e:
        print("\n\nERROR: There was an error saving the excel workbook. It may not exist.")
        logging.exception(e)

    if not loopSuccess:
        print("No emails were found. Please check the folder name variable and make sure it matches the one in outlook\n")

def main():
    # Connect to Outlook and process emails
    account = authenticate_account() # Function to authenticate with Outlook

    schedule.every().day.at("08:15").do(driver(account))

    while 1:
        schedule.run_pending()
        time.sleep(1)
    
if __name__ == "__main__":
    main()