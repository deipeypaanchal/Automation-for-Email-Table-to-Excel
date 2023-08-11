# Authors: Deipey Paanchal and Brandon Pyle

import imaplib
import email
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup

# Email credentials and server settings
IMAP_SERVER = 'imap.server.com'
EMAIL_ADDRESS = 'enter your email address'
PASSWORD = 'enter your app password'

# Sender's email address
SENDER_EMAIL = 'enter sender's email'

# Path to the master Excel workbook
MASTER_WORKBOOK_PATH = 'enter file path'

class EquipmentEntry:
    def __init__(self, site, equipment, message, last_state_change):
        self.site = site
        self.equipment = equipment
        self.message = message
        self.last_state_change = last_state_change


def parse_html_email(html_content, worksheet):
    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.findAll('table')

    if not table[0]:
        print("Error, no tables found in the email")
        return None

    if worksheet == 1:
        table_data = []
        for row in table[0].find_all('tr'):
            row_data = [cell.text.strip() for cell in row.find_all('td')]
            if row_data:
                table_data.append(row_data)
        return table_data
    else:
        email_table_data = []
        for row in table[0].find_all('tr')[1:]:
            columns = row.find_all('td')
            site, equipment, message, last_state_change = (col.text.strip() for col in columns)
            email_table_data.append(EquipmentEntry(site, equipment, message, last_state_change))

        return email_table_data

def append_table_data_to_worksheet1(table_data, worksheet1, added_rows):
    if table_data:
        for row_data in table_data[1:]:
            site, equipment, message, last_state_change = row_data[0], row_data[1], row_data[2], row_data[3]
            if (site, equipment, message, last_state_change) not in added_rows:
                added_rows.add((site, equipment, message, last_state_change))
                # Check if any of the values in the specific columns are not in the worksheet
                if not any(
                    (site, equipment, message, last_state_change) == (ws_site, ws_equipment, ws_message, ws_last_state_change)
                    for ws_site, ws_equipment, ws_message, ws_last_state_change in worksheet1.iter_rows(
                        values_only=True, min_col=1, max_col=4
                    )
                ):
                    worksheet1.append(row_data)

def update_table_data(existing_entries, new_entries, worksheet3):
    for new_entry in new_entries:
        for existing_entry in existing_entries:
            if new_entry.equipment == existing_entry.equipment and new_entry.site == existing_entry.site:
                existing_entry.message = new_entry.message
                existing_entry.last_state_change = new_entry.last_state_change
                break
        else:
            existing_entries.append(new_entry)

    # After updating the data in existing_entries, write it back to worksheet3
    worksheet3.delete_rows(2, worksheet3.max_row)  # Delete existing data (excluding header row)

    # Write the updated data to worksheet3
    for entry in existing_entries:
        worksheet3.append([entry.site, entry.equipment, entry.message, entry.last_state_change])

def read_existing_data_from_excel(worksheet3):
    try:
        data = list(worksheet3.values)  # Convert the worksheet data to a list of lists
        df = pd.DataFrame(data[1:], columns=data[0])  # Skip the header row
        existing_entries = [EquipmentEntry(row["Site"], row["Equipment"], row["Message"], row["Last State Change"]) for _, row in df.iterrows()]
    except:
        existing_entries = []

    return existing_entries

def main():
    # Connect to the email server
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_ADDRESS, PASSWORD)
    mail.select('inbox')

    # Search for emails from the specific sender
    status, data = mail.search(None, f'(FROM "{SENDER_EMAIL}")')
    if status != 'OK':
        print("Error occurred while searching for emails.")
        return

    email_ids = data[0].split()
    if not email_ids:
        print(f"No email found from the sender {SENDER_EMAIL}.")
        mail.close()
        mail.logout()
        return

    # Open the master workbook and access the sheets
    workbook = openpyxl.load_workbook(MASTER_WORKBOOK_PATH)
    worksheet1 = workbook["Sheet1"]
    worksheet2 = workbook["Sheet2"]
    worksheet3 = workbook["Sheet3"]

    # Keep track of added rows to avoid duplicates using a set
    added_rows = set(tuple(row) for row in worksheet1.iter_rows(values_only=True))

    existing_entries = read_existing_data_from_excel(worksheet3)

    for email_id in email_ids:
        status, data = mail.fetch(email_id, '(RFC822)')
        if status != 'OK':
            continue

        raw_email = data[0][1]
        email_message = email.message_from_bytes(raw_email)

        # Assuming the table is in the HTML body of the email
        for part in email_message.walk():
            if part.get_content_type() == 'text/html':
                html_body = part.get_payload(decode=True).decode('utf-8')

                # Extract the table data from the HTML body
                table_data = parse_html_table(html_body)

                # Append the table data to worksheet1, excluding duplicates
                append_table_data_to_worksheet1(table_data, worksheet1, added_rows)

                new_entries = parse_html_email(html_body)
                if new_entries:
                    update_table_data(existing_entries, new_entries, worksheet3)

    # Convert the combined data to a Pandas DataFrame
    data_dict = {
        "Site": [entry.site for entry in existing_entries],
        "Equipment": [entry.equipment for entry in existing_entries],
        "Message": [entry.message for entry in existing_entries],
        "Last State Change": [entry.last_state_change for entry in existing_entries]
    }
    df = pd.DataFrame(data_dict)

    # Save the data directly to the "Sheet3" of the master workbook using Openpyxl
    worksheet3.delete_rows(2, worksheet3.max_row)  # Delete existing data (excluding header row)
    for _, row in df.iterrows():
        worksheet3.append(row.tolist())

    # Find all rows with "not working" in the "Message" column
    not_working_rows = set()
    for row in worksheet3.iter_rows(values_only=True):
        if row[2] == "not working":  # Assuming "Message" is the 3rd column (index 2)
            not_working_rows.add(tuple(row))

    # Clear existing data in Sheet2
    worksheet2.delete_rows(2, worksheet2.max_row)

    # Append the not_working_rows to Sheet2
    for row_data in not_working_rows:
        worksheet2.append(row_data)


    # Save the modified workbook
    workbook.save(MASTER_WORKBOOK_PATH)
    workbook.close()

    print("Table data added/updated to the master worksheet.")

    # Disconnect from the email server
    mail.close()
    mail.logout()

if __name__ == "__main__":
    main()
